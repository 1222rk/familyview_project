# familyview_project/management/commands/populate_movies.py

import os
from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from django.conf import settings
from django.db import transaction
from familyview_project.models import Movie

class Command(BaseCommand):
    help = 'Load movie data from Excel file'

    def handle(self, *args, **options):
        file_path = os.path.join(settings.BASE_DIR, 'familyview_project', 'data', 'fyp_dataset.xlsx')
        wb = load_workbook(filename=file_path)
        ws = wb.active

        with transaction.atomic():
            row_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 5:
                    continue

                title, genre, release_date, age_rating, description = row[:5]

                if not title:
                    continue

                # Normalize age_rating
                valid_ratings = ["U", "PG", "12"]
                if age_rating not in valid_ratings:
                    age_rating = "12"

                # Extract just the first four digits for release_date
                if release_date is None:
                    year_str = ''
                else:
                    # Excel may give you a datetime/date or a float/int
                    if hasattr(release_date, 'year'):
                        year_str = str(release_date.year)
                    else:
                        # e.g. '2008.0' → '2008'
                        year_str = str(release_date)[:4]

                Movie.objects.update_or_create(
                    title=title,
                    defaults={
                        'genre': genre or '',
                        'release_date': year_str,
                        'age_rating': age_rating,
                        'description': description or '',
                    }
                )
                row_count += 1

        print(f'Data import complete. Processed {row_count} rows.')